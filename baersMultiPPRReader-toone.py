
from openpyxl import Workbook

import re




baersRegex = re.compile(r'''(\d{9})\s(\w{4})\s(.{30})\s(.{7})\s(\d{2})\s(.{9})''')

PPRQty = 2



wb = Workbook()

ws = wb.active

ws['A1'] = 'ITM_CD'
ws['B1'] = 'VE_CD'
ws['C1'] = 'VSN'
ws['D1'] = 'ADV_PRC'
ws['E1'] = 'ST'
ws['F1'] = 'CHNG_DATE'

ws['C5'] = "you didn't copy anything"

lines = []

ppr = ''

for pprnum in range(PPRQty):

    #read in ifr text file
    f = open('ppr{}.txt'.format(pprnum), 'r')
    
    for line in f:
        ppr += line
        
    f.close()


lines = baersRegex.findall(ppr)

rowcnt = 2

for line in range(len(lines)):
    ITM_CD, VE_CD, VSN, ADV_PRC, ST, CHNG_DATE = lines[line]



    print(VE_CD + ' ' + VSN)

    ws['A' + str(rowcnt)] = ITM_CD
    ws['B' + str(rowcnt)] = VE_CD
    ws['C' + str(rowcnt)] = VSN
    ws['D' + str(rowcnt)] = ADV_PRC
    ws['E' + str(rowcnt)] = ST
    ws['F' + str(rowcnt)] = CHNG_DATE


    rowcnt += 1


wb.save("PPR{}.xlsx".format(pprnum))
