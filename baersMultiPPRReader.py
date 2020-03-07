
from openpyxl import Workbook

import re

baersRegex = re.compile(r'''(\d{9})\s(\w{4})\s(.{30})\s(.{7})\s(.{7})\s(.{0,7})\s(\d{2})''')

PPRQty = 8

for pprnum in range(PPRQty):

    ppr = ''

    #read in ifr text file
    f = open('ppr{}.txt'.format(pprnum), 'r')
    
    for line in f:
        ppr += line
        
    f.close()


    lines = []

    lines = baersRegex.findall(ppr)

    wb = Workbook()

    ws = wb.active

    ws['A1'] = 'ITM_CD'
    ws['B1'] = 'VE_CD'
    ws['C1'] = 'VSN'
    ws['D1'] = 'ADV_PRC'
    ws['E1'] = 'RET_PRC'
    ws['F1'] = 'PRC1'
    ws['G1'] = 'ST'



    ws['C5'] = "you didn't copy amything"



    rowcnt = 2

    for line in range(len(lines)):
        ITM_CD, VE_CD, VSN, ST, ADV_PRC, RET_PRC, PRC1 = lines[line]

        #print(skus[item])
        #if str(vend) in carey_codes:

        print(VE_CD)

        ws['A' + str(rowcnt)] = ITM_CD
        ws['B' + str(rowcnt)] = VE_CD
        ws['C' + str(rowcnt)] = VSN
        ws['D' + str(rowcnt)] = ST
        ws['E' + str(rowcnt)] = (ADV_PRC)
        ws['F' + str(rowcnt)] = (RET_PRC)
        ws['G' + str(rowcnt)] = (PRC1)


        rowcnt += 1


    wb.save("PPR{}.xlsx".format(pprnum))
