
from openpyxl import Workbook

import re


#sbs = sales by store by sku reports 

baersRegex = re.compile(r'''(\d{2}-\w{3}-\d{2})\s+(.{2})\s+(.{9})\s+(.{3})\s+''')

SBSQty = 31



wb = Workbook()

ws = wb.active

ws['A1'] = 'DATE'
ws['B1'] = 'SO'
ws['C1'] = 'ITM_CD'
ws['D1'] = 'QTY'

ws['C5'] = "you didn't copy anything"

lines = []

sbs = ''

for sbsnum in range(SBSQty):

    #read in 
    f = open('sbs{}.txt'.format(sbsnum), 'r')
    
    for line in f:
        sbs += line
        
    f.close()


lines = baersRegex.findall(sbs)

rowcnt = 2

for line in range(len(lines)):
    DATE, SO, ITM_CD, QTY = lines[line]



    #print(DATE + ' ' + SO + ' ' + ITM_CD + ' ' + QTY)

    ws['A' + str(rowcnt)] = DATE
    ws['B' + str(rowcnt)] = SO
    ws['C' + str(rowcnt)] = ITM_CD
    ws['D' + str(rowcnt)] = QTY

    rowcnt += 1


wb.save("SBS_0thru{}.xlsx".format(sbsnum))
