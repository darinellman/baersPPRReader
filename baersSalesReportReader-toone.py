
from openpyxl import Workbook

import re


#sbs = sales by store by sku reports 

baersRegex = re.compile(r'''(\d{2})\s+(\d{9})\s+(\d+)''')

SBSQty = 12



wb = Workbook()

ws = wb.active

ws['A1'] = 'SO'
ws['B1'] = 'ITM_CD'
ws['C1'] = 'QTY'

ws['C5'] = "you didn't copy anything"

lines = []

sbs = ''

for sbsnum in range(SBSQty):

    #read in 
    f = open('sbs{}.txt'.format(sbsnum), 'r')
    
    for line in f:
        sbs += line
        
    f.close()


lines = baersRegex.findall(sbs)

rowcnt = 2

for line in range(len(lines)):
    SO, ITM_CD, QTY = lines[line]



    print(SO + ' ' + ITM_CD + ' ' + QTY)

    ws['A' + str(rowcnt)] = SO
    ws['B' + str(rowcnt)] = ITM_CD
    ws['C' + str(rowcnt)] = QTY


    rowcnt += 1


wb.save("SBS_0thru{}.xlsx".format(sbsnum))
